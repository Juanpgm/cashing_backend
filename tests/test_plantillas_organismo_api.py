"""API tests for the clone-oriented plantillas-organismo surface (docx clone,
phase 2): GET/POST /contratos/{id}/plantillas-organismo, GET /cuentas-cobro/
{id}/formato-valores and the clone-only cuenta-cobro.docx error path.
"""

from __future__ import annotations

from datetime import date
from typing import Any
from unittest.mock import AsyncMock

import pytest
from app.models.contrato import Contrato
from app.models.cuenta_cobro import CuentaCobro, EstadoCuentaCobro
from app.models.documento_fuente import DocumentoFuente, TipoDocumentoFuente
from app.models.plantilla_organismo import PlantillaOrganismo
from app.schemas.plantilla_organismo import EstructuraPlantillaLLM
from app.services import requisito_inference_service as svc
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

_CAMPOS = [
    {
        "direccion": "P0",
        "etiqueta": "Contrato No.",
        "campo": "contrato.numero_contrato",
        "valor_ejemplo": "ABC-123",
        "modo": "substring",
    },
    {
        "direccion": "T0.R0.C1",
        "etiqueta": "Valor Cuota a cancelar",
        "campo": "cuota.valor",
        "valor_ejemplo": "$ 4.800.000",
        "modo": "cell",
    },
]


@pytest.fixture
async def contrato(db: AsyncSession, test_user: dict[str, Any]) -> Contrato:
    user = test_user["user"]
    c = Contrato(
        usuario_id=user.id,
        numero_contrato="CTR-API-PO-001",
        objeto="Servicios profesionales",
        valor_total=12_000_000,
        valor_mensual=1_000_000,
        fecha_inicio=date(2024, 1, 1),
        fecha_fin=date(2024, 12, 31),
        entidad="COEMPRESAR",
    )
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return c


@pytest.fixture
async def cuenta(db: AsyncSession, contrato: Contrato) -> CuentaCobro:
    cc = CuentaCobro(
        contrato_id=contrato.id,
        mes=4,
        anio=2024,
        estado=EstadoCuentaCobro.BORRADOR,
        valor=1_000_000,
        numero_cuota=1,
    )
    db.add(cc)
    await db.commit()
    await db.refresh(cc)
    return cc


@pytest.fixture
async def plantilla(db: AsyncSession, contrato: Contrato) -> PlantillaOrganismo:
    p = PlantillaOrganismo(
        usuario_id=contrato.usuario_id,
        entidad="COEMPRESAR",
        entidad_normalizada="coempresar",
        tipo_documento="cuenta_cobro",
        formato="docx",
        estructura_json={
            "columnas": [],
            "secciones": [],
            "anexo_refs": [],
            "notas": "",
            "campos": _CAMPOS,
            "clonable": True,
            "avisos": [],
        },
    )
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


# ── GET /contratos/{id}/plantillas-organismo ─────────────────────────────────


async def test_listar_plantillas_organismo_agrupa_campos_por_scope(
    client: AsyncClient, test_user: dict[str, Any], contrato: Contrato, plantilla: PlantillaOrganismo
) -> None:
    resp = await client.get(f"/api/v1/contratos/{contrato.id}/plantillas-organismo/", headers=test_user["headers"])
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1
    item = data[0]
    assert item["tipo_documento"] == "cuenta_cobro"
    assert item["clonable"] is True
    assert item["campos"]["contrato"] == [
        {"campo": "contrato.numero_contrato", "etiqueta": "Contrato No.", "modo": "substring"}
    ]
    assert item["campos"]["cuota"][0]["campo"] == "cuota.valor"


async def test_listar_plantillas_organismo_vacio_sin_ingesta(
    client: AsyncClient, test_user: dict[str, Any], contrato: Contrato
) -> None:
    resp = await client.get(f"/api/v1/contratos/{contrato.id}/plantillas-organismo/", headers=test_user["headers"])
    assert resp.status_code == 200
    assert resp.json() == []


# ── POST /contratos/{id}/plantillas-organismo ────────────────────────────────


async def test_post_ingesta_devuelve_plantilla_clonable(
    client: AsyncClient,
    db: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    test_user: dict[str, Any],
    contrato: Contrato,
) -> None:
    doc = DocumentoFuente(
        usuario_id=contrato.usuario_id,
        contrato_id=contrato.id,
        storage_key="documentos/cuenta-cobro.docx",
        nombre="cuenta-cobro.docx",
        tipo=TipoDocumentoFuente.PLANTILLA,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)

    monkeypatch.setattr("app.adapters.storage.get_storage", lambda *_a, **_k: AsyncMock())
    monkeypatch.setattr(
        svc, "inferir_estructura_plantilla", AsyncMock(return_value=EstructuraPlantillaLLM(notas="carta"))
    )
    monkeypatch.setattr(svc, "detectar_campos_plantilla", AsyncMock(return_value=(_CAMPOS, ["Se descartó uno."])))

    resp = await client.post(
        f"/api/v1/contratos/{contrato.id}/plantillas-organismo/",
        headers=test_user["headers"],
        json={"documento_fuente_id": str(doc.id)},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["id"] is not None
    assert data["tipo_documento"] == "cuenta_cobro"
    assert data["clonable"] is True
    assert data["avisos"] == ["Se descartó uno."]
    assert set(data["campos"]) == {"contrato", "cuota"}


async def test_post_ingesta_degradada_devuelve_clonable_false(
    client: AsyncClient,
    db: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    test_user: dict[str, Any],
    contrato: Contrato,
) -> None:
    doc = DocumentoFuente(
        usuario_id=contrato.usuario_id,
        contrato_id=contrato.id,
        storage_key="documentos/ilegible.docx",
        nombre="ilegible.docx",
        tipo=TipoDocumentoFuente.PLANTILLA,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)

    monkeypatch.setattr("app.adapters.storage.get_storage", lambda *_a, **_k: AsyncMock())
    monkeypatch.setattr(svc, "inferir_estructura_plantilla", AsyncMock(return_value=None))

    resp = await client.post(
        f"/api/v1/contratos/{contrato.id}/plantillas-organismo/",
        headers=test_user["headers"],
        json={"documento_fuente_id": str(doc.id)},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["id"] is None
    assert data["clonable"] is False
    assert data["avisos"]


# ── GET /cuentas-cobro/{id}/formato-valores ──────────────────────────────────


async def test_formato_valores_devuelve_valores_y_aviso_sin_plantilla(
    client: AsyncClient, test_user: dict[str, Any], contrato: Contrato, cuenta: CuentaCobro
) -> None:
    resp = await client.get(f"/api/v1/cuentas-cobro/{cuenta.id}/formato-valores", headers=test_user["headers"])
    assert resp.status_code == 200
    data = resp.json()
    assert data["valores"]["cuota.valor"] == "$ 1.000.000"
    assert data["valores"]["cuota.periodo"] == "Abril de 2024"
    assert data["valores"]["computed.cuota_ordinal_letras"] == "primera"
    assert len(data["avisos"]) == 1


async def test_formato_valores_avisa_campos_sin_dato_y_checkbox_indeterminado(
    client: AsyncClient,
    db: AsyncSession,
    test_user: dict[str, Any],
    contrato: Contrato,
    cuenta: CuentaCobro,
) -> None:
    """RESI-001/RELI-001: a clonable plantilla with campos that resolve to no
    value must be surfaced in avisos (naming the etiquetas), and a checkbox
    whose option texts can't be read (no fuente bytes) must be omitted from
    valores with the manual-marking aviso."""
    db.add(
        PlantillaOrganismo(
            usuario_id=contrato.usuario_id,
            entidad="COEMPRESAR",
            entidad_normalizada="coempresar",
            tipo_documento="cuenta_cobro",
            formato="docx",
            estructura_json={
                "columnas": [],
                "secciones": [],
                "anexo_refs": [],
                "notas": "",
                "campos": [
                    *_CAMPOS,
                    {
                        "direccion": "T1.R0.C1",
                        "etiqueta": "Supervisor",
                        "campo": "contrato.supervisor_nombre",
                        "valor_ejemplo": "Pepito Ejemplo",
                        "modo": "cell",
                    },
                    {
                        "direccion": "T1.R1.C1",
                        "etiqueta": "Dependencia",
                        "campo": "contrato.dependencia",
                        "valor_ejemplo": "Dependencia Ejemplo",
                        "modo": "cell",
                    },
                    {
                        "direccion": "T2.R0.C0",
                        "etiqueta": "Tipo de informe",
                        "campo": "computed.tipo_informe",
                        "valor_ejemplo": "X",
                        "modo": "checkbox",
                        "direccion_alternativa": "T2.R0.C1",
                        "valor_alternativo": "X",
                    },
                ],
                "clonable": True,
                "avisos": [],
            },
            # No fuente_documento_id → checkbox option texts unreadable.
        )
    )
    await db.commit()

    resp = await client.get(f"/api/v1/cuentas-cobro/{cuenta.id}/formato-valores", headers=test_user["headers"])
    assert resp.status_code == 200
    data = resp.json()
    # Resolvable campos still preview normally.
    assert data["valores"]["cuota.valor"] == "$ 1.000.000"
    # The unreadable checkbox is omitted, never guessed.
    assert "computed.tipo_informe" not in data["valores"]
    aviso_blanco = next(a for a in data["avisos"] if "quedarán en blanco" in a)
    assert "Supervisor" in aviso_blanco and "Dependencia" in aviso_blanco
    assert "Contrato No." not in aviso_blanco  # resolved campo not listed
    assert any("marcar manualmente" in a for a in data["avisos"])


# ── GET /cuentas-cobro/{id}/cuenta-cobro.docx — clone-only error path ────────


async def test_cuenta_cobro_docx_sin_plantilla_es_422(
    client: AsyncClient, test_user: dict[str, Any], contrato: Contrato, cuenta: CuentaCobro
) -> None:
    resp = await client.get(f"/api/v1/cuentas-cobro/{cuenta.id}/cuenta-cobro.docx", headers=test_user["headers"])
    assert resp.status_code == 422
    assert "cuenta de cobro" in resp.json()["detail"]


# ── GET /cuentas-cobro/{id}/documento-soporte.xlsx (xlsx clone, phase 4) ─────


def _build_xlsx() -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Documento Soporte"
    ws["B2"] = "Valor"
    ws["C2"] = "$ 999"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_documento_soporte_xlsx_sin_plantilla_es_422(
    client: AsyncClient, test_user: dict[str, Any], contrato: Contrato, cuenta: CuentaCobro
) -> None:
    resp = await client.get(f"/api/v1/cuentas-cobro/{cuenta.id}/documento-soporte.xlsx", headers=test_user["headers"])
    assert resp.status_code == 422
    assert "documento soporte" in resp.json()["detail"]


async def test_documento_soporte_xlsx_descarga_clon_rellenado(
    client: AsyncClient,
    db: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    test_user: dict[str, Any],
    contrato: Contrato,
    cuenta: CuentaCobro,
) -> None:
    from io import BytesIO

    from app.services import informe_service
    from openpyxl import load_workbook

    doc = DocumentoFuente(
        usuario_id=contrato.usuario_id,
        contrato_id=contrato.id,
        storage_key="documentos/documento-soporte.xlsx",
        nombre="documento-soporte.xlsx",
        tipo=TipoDocumentoFuente.PLANTILLA,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    db.add(
        PlantillaOrganismo(
            usuario_id=contrato.usuario_id,
            entidad="COEMPRESAR",
            entidad_normalizada="coempresar",
            tipo_documento="documento_soporte",
            formato="xlsx",
            estructura_json={
                "columnas": [],
                "secciones": [],
                "anexo_refs": [],
                "notas": "",
                "campos": [
                    {
                        "direccion": "Documento Soporte!C2",
                        "etiqueta": "Valor",
                        "campo": "cuota.valor",
                        "valor_ejemplo": "$ 999",
                        "modo": "cell",
                    }
                ],
                "clonable": True,
                "avisos": [],
            },
            fuente_documento_id=doc.id,
        )
    )
    await db.commit()

    storage = AsyncMock()
    storage.download = AsyncMock(return_value=_build_xlsx())
    monkeypatch.setattr(informe_service, "_get_storage", lambda *_a, **_k: storage)

    resp = await client.get(f"/api/v1/cuentas-cobro/{cuenta.id}/documento-soporte.xlsx", headers=test_user["headers"])
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert resp.headers["x-es-borrador"] == "true"
    assert 'filename="documento-soporte-CTR-API-PO-001-2024-04.xlsx"' in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert wb["Documento Soporte"]["C2"].value == "$ 1.000.000"  # cuota.valor filled
    assert wb["Documento Soporte"]["B2"].value == "Valor"  # untouched label
