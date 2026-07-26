"""Tests for xlsx_clone_service pure functions (skeleton, validation, filling).

Builds a small in-memory workbook with openpyxl — two sheets with labeled
cells plus a currency-formatted cell — and round-trips it through
extraer_esqueleto_xlsx / validar_campos_xlsx / rellenar_xlsx.
"""

from __future__ import annotations

from io import BytesIO

from app.services.xlsx_clone_service import (
    extraer_esqueleto_xlsx,
    rellenar_xlsx,
    validar_campos_xlsx,
)
from openpyxl import Workbook, load_workbook

CAMPOS = [
    {
        "direccion": "Documento Soporte!C2",
        "etiqueta": "Contrato No.",
        "campo": "contrato.numero_contrato",
        "valor_ejemplo": "ABC-123",
        "modo": "cell",
    },
    {
        "direccion": "Anexo!B1",
        "etiqueta": "Valor cuota",
        "campo": "cuota.valor",
        "valor_ejemplo": "$ 4.800.000",
        "modo": "cell",
    },
]

CAMPO_BOGUS = {
    "direccion": "Fantasma!Z9",
    "etiqueta": "Fantasma",
    "campo": "fantasma",
    "valor_ejemplo": "nada",
    "modo": "cell",
}


def _build_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Documento Soporte"
    ws["A1"] = "DISTRITO DE SANTIAGO DE CALI"
    ws["B2"] = "Contrato No."
    ws["C2"] = "ABC-123"
    ws["C3"] = 4_800_000
    ws["C3"].number_format = '"$"#,##0'
    ws2 = wb.create_sheet("Anexo")
    ws2["A1"] = "Valor cuota"
    ws2["B1"] = "$ 4.800.000"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExtraerEsqueletoXlsx:
    def test_emits_addressed_lines_per_sheet(self) -> None:
        esq = extraer_esqueleto_xlsx(_build_xlsx())
        lines = esq.splitlines()
        assert "Documento Soporte!A1: DISTRITO DE SANTIAGO DE CALI" in lines
        assert "Documento Soporte!C2: ABC-123" in lines
        assert "Documento Soporte!C3: 4800000" in lines  # stored value, not display format
        assert "Anexo!B1: $ 4.800.000" in lines
        assert lines.index("Documento Soporte!A1: DISTRITO DE SANTIAGO DE CALI") < lines.index("Anexo!B1: $ 4.800.000")

    def test_truncates_at_line_boundary_with_marker(self) -> None:
        esq = extraer_esqueleto_xlsx(_build_xlsx(), max_chars=50)
        lines = esq.splitlines()
        assert lines[-1] == "... [esqueleto truncado]"
        assert all(len(line) <= 50 for line in lines[:-1])


class TestValidarCamposXlsx:
    def test_keeps_resolvable_campos_and_drops_bogus_with_aviso(self) -> None:
        validos, avisos = validar_campos_xlsx(_build_xlsx(), [*CAMPOS, CAMPO_BOGUS])
        assert validos == CAMPOS
        assert len(avisos) == 1
        assert "Fantasma" in avisos[0]
        assert "Fantasma!Z9" in avisos[0]

    def test_drops_campo_whose_valor_ejemplo_is_absent(self) -> None:
        campo = {**CAMPOS[0], "valor_ejemplo": "ZZZ-999"}
        validos, avisos = validar_campos_xlsx(_build_xlsx(), [campo])
        assert validos == []
        assert len(avisos) == 1

    def test_drops_non_cell_modo_with_aviso(self) -> None:
        campo = {**CAMPOS[0], "modo": "substring"}
        validos, avisos = validar_campos_xlsx(_build_xlsx(), [campo])
        assert validos == []
        assert len(avisos) == 1
        assert "substring" in avisos[0]


class TestRellenarXlsx:
    def test_writes_values_and_preserves_number_format(self) -> None:
        valores = {
            "contrato.numero_contrato": "NEW-999",
            # "cuota.valor" intentionally absent → skipped silently
        }
        out = load_workbook(BytesIO(rellenar_xlsx(_build_xlsx(), CAMPOS, valores)))

        ws = out["Documento Soporte"]
        assert ws["C2"].value == "NEW-999"
        # campo without a value is untouched
        assert out["Anexo"]["B1"].value == "$ 4.800.000"
        # untouched cells keep their content and currency format after the round trip
        assert ws["C3"].value == 4_800_000
        assert ws["C3"].number_format == '"$"#,##0'
        assert ws["A1"].value == "DISTRITO DE SANTIAGO DE CALI"

    def test_unresolvable_direccion_is_skipped_silently(self) -> None:
        out = load_workbook(BytesIO(rellenar_xlsx(_build_xlsx(), [CAMPO_BOGUS], {"fantasma": "boo"})))
        assert out["Documento Soporte"]["C2"].value == "ABC-123"
