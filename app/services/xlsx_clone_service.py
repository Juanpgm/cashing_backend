"""Xlsx clone service — pure functions to clone-and-fill entity Excel templates.

Twin of `docx_clone_service` for the Documento Soporte spreadsheet family:
edits the original workbook in place (via openpyxl) so formatting, formulas
and layout survive. No DB, no LLM — callers pass xlsx bytes plus a list of
campo dicts and get bytes back.

Address scheme ("direccion"): ``{sheet_name}!{cell}`` over visible sheets,
e.g. ``Documento Soporte!C7``. Only modo ``cell`` is supported — a spreadsheet
cell IS the value; there is no run fragmentation to work around.

Campo dict contract (produced upstream, consumed here — same shape as the
docx twin):
    {"direccion": "Documento Soporte!C7", "etiqueta": "Contrato No.",
     "campo": "contrato.numero_contrato", "valor_ejemplo": "ABC-123",
     "modo": "cell"}
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import coordinate_to_tuple

from app.core.text_match import normalize

_TRUNCATION_MARKER = "... [esqueleto truncado]"


def _norm(text: str | None) -> str:
    """Accent/case-insensitive, whitespace-collapsed form for tolerant matching."""
    return " ".join(normalize(text).split())


def _clean(value: object) -> str:
    """Stored cell value (formula string, number, text) as one collapsed line."""
    return " ".join(str(value).split())


def extraer_esqueleto_xlsx(xlsx_bytes: bytes, max_chars: int = 14_000) -> str:
    """Render visible sheets as one ``{sheet}!{cell}: text`` line per non-empty cell.

    Formula cells emit their stored value as-is (data_only=False). Output is cut
    at a line boundary once it would exceed ``max_chars``, appending a truncation
    marker line — same convention as `docx_clone_service.extraer_esqueleto`.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    lines: list[str] = []
    total = 0
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = _clean(cell.value)
                if not text:
                    continue
                line = f"{ws.title}!{cell.coordinate}: {text}"
                if total + len(line) + 1 > max_chars:
                    lines.append(_TRUNCATION_MARKER)
                    return "\n".join(lines)
                lines.append(line)
                total += len(line) + 1
    return "\n".join(lines)


def _resolver_celda(wb: Workbook, direccion: str) -> Cell | None:
    """Resolve a ``{sheet}!{cell}`` address; None on malformed/unknown/out-of-bounds."""
    try:
        direccion = (direccion or "").strip()
        sheet_name, sep, coord = direccion.rpartition("!")
        sheet_name = sheet_name.strip("'")  # tolerate quoted names ('Documento Soporte'!C7)
        if not sep or sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        fila, columna = coordinate_to_tuple(coord.strip().upper())
        if not (1 <= fila <= ws.max_row and 1 <= columna <= ws.max_column):
            return None
        return ws.cell(row=fila, column=columna)
    except Exception:  # defensive: address input is untrusted (LLM-produced)
        return None


def validar_campos_xlsx(xlsx_bytes: bytes, campos: list[dict]) -> tuple[list[dict], list[str]]:
    """Keep only modo="cell" campos whose direccion resolves and whose
    valor_ejemplo appears in that cell's text.

    Returns ``(campos_validos, avisos)`` — avisos are user-facing Spanish strings.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    validos: list[dict] = []
    avisos: list[str] = []
    for campo in campos:
        etiqueta = campo.get("etiqueta") or campo.get("campo") or "?"
        modo = campo.get("modo", "cell")
        if modo != "cell":
            avisos.append(f"Se descartó el campo '{etiqueta}': el modo '{modo}' no está soportado en plantillas xlsx.")
            continue
        direccion = campo.get("direccion", "")
        cell = _resolver_celda(wb, direccion)
        if cell is None:
            avisos.append(f"Se descartó el campo '{etiqueta}': la dirección '{direccion}' no existe en la plantilla.")
            continue
        valor_ejemplo = campo.get("valor_ejemplo", "")
        if _norm(valor_ejemplo) not in _norm(_clean(cell.value) if cell.value is not None else ""):
            avisos.append(
                f"Se descartó el campo '{etiqueta}': el valor de ejemplo '{valor_ejemplo}' no aparece en '{direccion}'."
            )
            continue
        validos.append(campo)
    return validos, avisos


def rellenar_xlsx(xlsx_bytes: bytes, campos: list[dict], valores: dict[str, str]) -> bytes:
    """Fill the template: write each campo's value directly into its cell.

    Campos without a value (or whose direccion no longer resolves, or which
    point at a read-only merged non-anchor cell) are skipped silently, matching
    `docx_clone_service.rellenar`. openpyxl preserves cell formatting natively
    on rewrite.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    for campo in campos:
        clave = campo.get("campo")
        if clave is None or clave not in valores:
            continue
        cell = _resolver_celda(wb, campo.get("direccion", ""))
        if cell is None or isinstance(cell, MergedCell):
            continue
        cell.value = valores[clave]
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
